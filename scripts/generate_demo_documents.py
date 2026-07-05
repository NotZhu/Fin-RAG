"""生成 demo-documents 企业级多格式业务文档"""

from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
import zipfile
from html import escape
from pathlib import Path
from xml.sax.saxutils import escape as xml_escape

import fitz
from PIL import Image, ImageDraw, ImageFont


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "output" / "demo-documents"
PAGE_SIZE = (1240, 1754)
FONT_CANDIDATES = [
    Path(r"C:\Windows\Fonts\msyh.ttc"),
    Path(r"C:\Windows\Fonts\simsun.ttc"),
    Path(r"C:\Windows\Fonts\arial.ttf"),
]
CREDIT_REPORT_TITLE = "华东智造集团 2026H1 授信审查报告"
EMU_PER_INCH = 914400
PPT_SLIDE_CX = 12192000
PPT_SLIDE_CY = 6858000


def parse_args() -> argparse.Namespace:
    """解析命令行参数。"""
    parser = argparse.ArgumentParser(description="生成 demo-documents 企业级 RAG 评测文档")
    parser.add_argument("--out-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--clean", action="store_true", help="生成前清空输出目录")
    return parser.parse_args()


def main() -> int:
    """生成文档并打印产物路径。"""
    args = parse_args()
    generated = generate_demo_documents(args.out_dir.resolve(), clean=args.clean)
    for path in generated:
        print(path)
    return 0


def generate_demo_documents(output_dir: Path, *, clean: bool = False) -> list[Path]:
    """生成全部 demo-documents 业务资料。"""
    if clean and output_dir.exists():
        shutil.rmtree(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    return [
        write_credit_review_pdf(output_dir / "credit_review_report.pdf"),
        write_accounts_receivable_xlsx(output_dir / "accounts_receivable_aging.xlsx"),
        write_procurement_contract_docx(output_dir / "procurement_contract.docx"),
        write_risk_limit_policy_html(output_dir / "risk_limit_policy.html"),
        write_collateral_inventory_csv(output_dir / "collateral_inventory.csv"),
        write_aml_due_diligence_txt(output_dir / "aml_enhanced_due_diligence.txt"),
        write_erp_snapshot_json(output_dir / "erp_financial_snapshot.json"),
        write_supplier_esg_pptx(output_dir / "supplier_esg_review.pptx"),
        write_board_minutes_md(output_dir / "board_financing_minutes.md"),
        write_bank_statement_scan(output_dir / "bank_statement_scan.png"),
    ]


def write_credit_review_pdf(path: Path) -> Path:
    """写入可抽取文本、结构更接近真实审查包的授信审查 PDF。"""
    doc = fitz.open()

    page = _new_pdf_page(doc, "报告编号：CR-2026-H1-017    客户经理：陆明    审查日期：2026-06-30")
    y = 96
    y = _pdf_text(page, "一、授信结论", 50, y, fontsize=14, bold=True)
    for paragraph in [
        "本报告用于华东智造集团 2026H1 综合授信复核，审查范围覆盖经营现金流、客户集中度、应收账款账龄、押品覆盖、采购合同约束和贷后触发条件。",
        "审查意见为维持综合授信额度 3.5 亿元，其中流动资金贷款 2.2 亿元、银行承兑汇票 0.8 亿元、保函 0.5 亿元；评级保持 A-，展望调整为关注。",
        "新增提款需满足两个条件：三季度经营现金流转正，逾期应收款占比降至 5% 以下。额度用途限定为订单备货、核心客户交付和税费周转，不得用于股权投资、关联拆借或非经营性支出。",
    ]:
        y = _pdf_text(page, paragraph, 64, y, fontsize=10.2)
    y = _pdf_table(
        page,
        "授信结构与审批约束",
        y + 4,
        ["品种", "额度", "期限", "提款约束"],
        [
            ["流动资金贷款", "2.2 亿元", "12 个月", "提款前核验订单、发票、出库和回款计划"],
            ["银行承兑汇票", "0.8 亿元", "6 个月", "保证金比例不低于 25%，不得循环开票"],
            ["履约保函", "0.5 亿元", "按项目匹配", "仅限新能源产线交付项目"],
        ],
        [96, 78, 82, 246],
    )
    y = _pdf_text(page, "核心判断", 50, y + 8, fontsize=12, bold=True)
    for paragraph in [
        "公司订单储备仍能支撑未来两个季度收入确认，但现金回收明显滞后，收入增长尚未同步转化为可用现金。",
        "授信维持的关键不在新增额度，而在提款节奏、核心客户回款闭环和贷后触发器执行。若现金流继续弱化，应暂停新增敞口并保留存量压降安排。",
    ]:
        y = _pdf_text(page, paragraph, 64, y, fontsize=10.2)

    page = _new_pdf_page(doc, "财务分析页：现金流、利润率、周转效率")
    y = 96
    y = _pdf_text(page, "二、经营现金流与偿债能力", 50, y, fontsize=14, bold=True)
    for paragraph in [
        "2026H1 经营现金流为 0.8 亿元，同比下降 -57.9%，主要原因是新能源客户回款周期延长、验收节点后移以及部分客户采用票据结算。",
        "营业收入 15.4 亿元，同比增长 20.3%；毛利率 29.6%，较上年同期下降 1.6 个百分点。收入增长质量弱于账面增速，需结合银行流水和 ERP 出库记录持续复核。",
    ]:
        y = _pdf_text(page, paragraph, 64, y, fontsize=10.2)
    y = _pdf_table(
        page,
        "财务指标矩阵",
        y + 4,
        ["指标", "2025H1", "2026H1", "变动", "审查含义"],
        [
            ["营业收入", "12.8 亿元", "15.4 亿元", "+20.3%", "订单增长真实，但回款不同步"],
            ["经营现金流", "1.9 亿元", "0.8 亿元", "-57.9%", "现金创造能力明显走弱"],
            ["毛利率", "31.2%", "29.6%", "-1.6pct", "原材料和交付成本挤压利润"],
            ["短债覆盖倍数", "1.34x", "1.18x", "-0.16x", "已低于 1.20x 管理线"],
            ["逾期应收款占比", "3.8%", "6.2%", "+2.4pct", "超过 5% 贷后关注阈值"],
            ["存货周转天数", "91 天", "104 天", "+13 天", "备货增加占用营运资金"],
        ],
        [82, 78, 78, 70, 194],
    )
    y = _pdf_text(page, "审查提示", 50, y + 8, fontsize=12, bold=True)
    for paragraph in [
        "短债覆盖倍数已接近提款限制线，应将新增提款与经营现金流转正、逾期应收款清收和库存库龄治理绑定。",
        "若三季度仍未形成现金回收改善，综合授信可维持但不得扩大敞口，同时要求客户经理按月核验银行流水摘要。",
    ]:
        y = _pdf_text(page, paragraph, 64, y, fontsize=10.2)

    page = _new_pdf_page(doc, "风险分析页：客户集中度、押品、压力测试")
    y = 96
    y = _pdf_text(page, "三、客户集中度与应收账款", 50, y, fontsize=14, bold=True)
    for paragraph in [
        "前五大客户收入占比 48.5%，超过 45% 的关注阈值；逾期应收款占比 6.2%，超过 5% 的管理线。延期回款主要集中于宁波新能源装备有限公司、合肥智能产线科技有限公司和上海储能装备有限公司。",
        "贷后检查需核验销售合同、发票、发货单、回款承诺和银行流水，避免收入确认与现金回收脱节。核心客户回款承诺未兑现时，应压缩新增赊销额度。",
    ]:
        y = _pdf_text(page, paragraph, 64, y, fontsize=10.2)
    y = _pdf_table(
        page,
        "押品与合同约束",
        y + 4,
        ["项目", "当前状态", "覆盖/比例", "核验动作"],
        [
            ["杭州厂区土地及厂房", "抵押顺位第一", "覆盖率 60%", "每年复评，关注权属限制"],
            ["数控加工中心设备组", "已办理抵押登记", "抵质押率 40%", "抽查设备清单和保险单"],
            ["核心客户应收账款池", "需月度核验", "抵质押率 50%", "核对发票、回款和买方确认"],
            ["采购合同付款条件", "20/50/30", "运行 60 天后尾款", "避免预付款形成资金挪用"],
        ],
        [116, 118, 84, 184],
    )
    y = _pdf_table(
        page,
        "压力测试情景",
        y + 8,
        ["情景", "假设", "短债覆盖倍数", "授信动作"],
        [
            ["基准", "三季度回款按计划到账", "1.18x", "维持额度，按订单提款"],
            ["中度压力", "核心客户回款延迟 30 天", "1.07x", "压缩提款频率，追加流水核验"],
            ["重度压力", "回款延迟 60 天且库存继续上升", "0.96x", "暂停新增敞口并保留提款节奏限制"],
        ],
        [78, 190, 92, 142],
    )

    page = _new_pdf_page(doc, "贷后管理页：触发器、责任人、复核节奏")
    y = 96
    y = _pdf_text(page, "四、采购合同与资金用途核验", 50, y, fontsize=14, bold=True)
    for paragraph in [
        "采购合同付款条件为预付款 20%、验收后支付 50%、稳定运行 60 天后支付 30%。新增提款不得直接覆盖长期资本性支出，需与订单备货、核心客户交付和税费周转逐笔匹配。",
        "若出现同日多笔拆分回款、异常大额预付款或新增对外担保，应触发增强尽调和授信复核，法务合规部需同步留存复核工作底稿。",
    ]:
        y = _pdf_text(page, paragraph, 64, y, fontsize=10.2)
    y = _pdf_text(page, "五、贷后监测安排", 50, y + 4, fontsize=14, bold=True)
    y = _pdf_table(
        page,
        "贷后触发器矩阵",
        y + 2,
        ["触发器", "阈值", "监测频率", "处置动作"],
        [
            ["短债覆盖倍数", "连续两个月低于 1.1x", "月度", "暂停新增提款，提交资金缺口测算"],
            ["逾期应收款占比", "未降至 5% 以下", "月度", "冻结新增赊销，建立回款专项台账"],
            ["客户集中度", "前五大客户超过 45%", "季度", "补充客户分散计划并上会审批"],
            ["异常资金流水", "同日拆分或大额预付款", "实时", "启动增强尽调，核验合同和发票"],
            ["对外担保", "新增或超过净资产 25%", "事件触发", "董事会三分之二以上通过后方可执行"],
        ],
        [118, 116, 72, 196],
    )
    y = _pdf_text(page, "六、审查意见", 50, y + 8, fontsize=14, bold=True)
    for paragraph in [
        "综合判断，公司订单充足但现金回收压力明显，授信维持需绑定回款进度、提款用途核验和押品动态复核。下一次复核时间为 2026-09-30。",
        "建议风险委员会同意维持存量综合授信，但将质量门禁写入贷后清单：经营现金流转正、逾期应收款压降、押品权属核验和异常流水复核。",
    ]:
        y = _pdf_text(page, paragraph, 64, y, fontsize=10.2)

    _pdf_footer(doc)
    doc.save(path, garbage=4, deflate=True)
    doc.close()
    return path


def write_accounts_receivable_xlsx(path: Path) -> Path:
    """写入应收账款账龄和回款计划工作簿。"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except Exception as exc:
        raise RuntimeError("生成 XLSX 样例需要安装 openpyxl") from exc

    workbook = Workbook()
    aging = workbook.active
    aging.title = "账龄明细"
    aging.append(["客户名称", "账龄区间", "应收余额", "逾期金额", "回款责任人"])
    for row in [
        ["宁波新能源装备有限公司", "90-180 天", "3,260 万元", "860 万元", "周岚"],
        ["苏州精密传动股份有限公司", "60-90 天", "1,780 万元", "320 万元", "陈昊"],
        ["合肥智能产线科技有限公司", "180 天以上", "940 万元", "610 万元", "王珂"],
        ["南京机器人集成有限公司", "30-60 天", "1,120 万元", "0 万元", "李青"],
        ["常州新能源电控有限公司", "90-180 天", "1,460 万元", "280 万元", "马原"],
        ["无锡精工自动化有限公司", "0-30 天", "760 万元", "0 万元", "郑蕾"],
        ["上海储能装备有限公司", "180 天以上", "520 万元", "410 万元", "周岚"],
        ["杭州工业软件有限公司", "60-90 天", "630 万元", "90 万元", "王珂"],
    ]:
        aging.append(row)

    plan = workbook.create_sheet("回款计划")
    plan.append(["客户名称", "承诺回款日", "计划金额", "风控动作", "跟踪频率"])
    for row in [
        ["宁波新能源装备有限公司", "2026-07-31", "500 万元", "纳入周度跟踪", "每周"],
        ["合肥智能产线科技有限公司", "2026-08-15", "300 万元", "暂停新增赊销", "每周"],
        ["上海储能装备有限公司", "2026-08-31", "260 万元", "要求补充回款确认函", "双周"],
        ["常州新能源电控有限公司", "2026-09-10", "200 万元", "压缩信用账期", "双周"],
    ]:
        plan.append(row)

    concentration = workbook.create_sheet("集中度分析")
    concentration.append(["指标", "当前值", "阈值", "状态", "处置建议"])
    for row in [
        ["前五大客户收入占比", "48.5%", "45%", "关注", "补充客户分散计划"],
        ["逾期应收款占比", "6.2%", "5%", "关注", "要求回款专项台账"],
        ["短债覆盖倍数", "1.18x", "1.20x", "轻微承压", "保留提款节奏限制"],
        ["存货周转天数", "104 天", "95 天", "关注", "提交库龄治理进展"],
    ]:
        concentration.append(row)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    border = Border(
        left=Side(style="thin", color="B7C7D6"),
        right=Side(style="thin", color="B7C7D6"),
        top=Side(style="thin", color="B7C7D6"),
        bottom=Side(style="thin", color="B7C7D6"),
    )
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.font = Font(name="Microsoft YaHei", bold=True, color="111827")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
        for index, width in enumerate([28, 18, 18, 18, 24], start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width
    workbook.save(path)
    return path


def write_procurement_contract_docx(path: Path) -> Path:
    """写入带章节和表格的采购合同 DOCX。"""
    sections = [
        ("采购合同", ["合同编号：PC-2026-0712", "甲方：华东智造集团，乙方：嘉兴精密部件有限公司。"]),
        ("一、采购背景", ["本合同用于华东智造集团新能源产线扩产项目，采购对象为伺服模组、控制板卡和关键装配组件。"]),
        ("二、采购清单", ["采购物料、规格、数量、交付节点和质量等级详见清单表。"]),
        ("三、付款条件", ["预付款 20%，到货验收合格后支付 50%，剩余 30% 在稳定运行 60 天后支付。"]),
        ("四、验收标准", ["关键零部件尺寸偏差不得超过 0.02mm，批次不良率不得高于 0.8%，供应商需提供出厂检测报告。"]),
        ("五、交付与物流", ["乙方负责运输、保险和入库前风险；甲方仓库完成外观检查后出具到货确认。"]),
        ("六、违约条款", ["逾期交付超过 10 个工作日，乙方按未交付金额每日 0.05% 支付违约金。"]),
        ("七、质量追索", ["质保期内出现批量质量问题，乙方应在 48 小时内响应，并承担返工、替换和停线损失。"]),
        ("八、争议解决", ["争议先由双方业务负责人协商；协商不成的，提交甲方所在地人民法院处理。"]),
    ]
    body = [
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">',
        "<w:body>",
    ]
    for title, paragraphs in sections:
        body.append(_docx_paragraph(title, style="Heading1"))
        body.extend(_docx_paragraph(paragraph) for paragraph in paragraphs)
        if title == "二、采购清单":
            body.append(
                _docx_table(
                    ["物料", "规格", "数量", "交付节点"],
                    [
                        ["伺服模组", "HM-SV-220", "120 套", "2026-08-20"],
                        ["控制板卡", "HM-CTRL-X7", "80 套", "2026-09-05"],
                        ["高精度导轨", "HG-45A", "160 根", "2026-09-12"],
                        ["视觉检测单元", "VS-900", "24 套", "2026-09-25"],
                    ],
                )
            )
        if title == "四、验收标准":
            body.append(
                _docx_table(
                    ["检查项", "标准", "抽检比例", "不合格处理"],
                    [
                        ["尺寸精度", "偏差 <= 0.02mm", "15%", "整批复检"],
                        ["运行稳定性", "连续运行 72 小时无异常", "100%", "暂停付款"],
                        ["外观包装", "无磕碰、无锈蚀", "20%", "退换货"],
                    ],
                )
            )
        if title == "七、质量追索":
            body.append(
                _docx_table(
                    ["场景", "响应时限", "责任承担", "证据材料"],
                    [
                        ["批量故障", "48 小时", "乙方承担返工和替换", "检测报告、现场照片"],
                        ["停线损失", "24 小时", "按实际停线损失协商赔付", "停线记录、维修单"],
                    ],
                )
            )
    body.append(
        '<w:sectPr><w:pgSz w:w="11906" w:h="16838"/><w:pgMar w:top="1134" w:right="1134" w:bottom="1134" w:left="1134"/></w:sectPr>'
    )
    body.extend(["</w:body>", "</w:document>"])
    _write_docx_package(path, "\n".join(body))
    return path


def write_risk_limit_policy_html(path: Path) -> Path:
    """写入行业风险限额 HTML 制度。"""
    rows = [
        ["新能源装备", "单户授信不超过 4 亿元", "客户集中度超过 45% 需上会审批"],
        ["智能制造", "流贷期限不超过 12 个月", "短债覆盖倍数低于 1.1x 暂停新增"],
        ["金属加工", "抵押覆盖率不低于 120%", "环保处罚未整改不得新增敞口"],
        ["工业软件", "年度经常性收入覆盖贷款本息 1.5 倍以上", "客户流失率超过 12% 需复核"],
    ]
    html = [
        "<!doctype html>",
        '<html lang="zh-CN">',
        "<head><meta charset=\"utf-8\"><title>行业风险限额政策</title>",
        "<style>body{font-family:Microsoft YaHei,Arial,sans-serif;line-height:1.7;margin:36px;color:#1f2937}table{border-collapse:collapse;width:100%;margin:12px 0 24px}th,td{border:1px solid #b8c2cc;padding:8px 10px;vertical-align:top}th{background:#e8f0f6}</style></head>",
        "<body>",
        "<h1>华东智造集团授信风险限额政策</h1>",
        "<section><h2>一、适用范围</h2><p>本政策用于贷前审查、额度核定、提款审核和贷后监测，覆盖行业限额、集中度阈值和审批例外。</p></section>",
        "<section><h2>二、行业限额矩阵</h2><table><thead><tr><th>行业</th><th>限额</th><th>审批例外</th></tr></thead><tbody>",
    ]
    for row in rows:
        html.append("<tr>" + "".join(f"<td>{escape(value)}</td>" for value in row) + "</tr>")
    html.extend(
        [
            "</tbody></table></section>",
            "<section><h2>三、贷后触发器</h2><p>短债覆盖倍数低于 1.1x、逾期应收款占比超过 5%、同日多笔拆分回款或新增对外担保均触发贷后复核。</p></section>",
            "<section><h2>四、例外审批材料</h2><p>例外审批需提交经营现金流预测、核心客户回款计划、押品价值复核、董事会授权文件和增强尽调结论。</p></section>",
            "<section><h2>五、报告频率</h2><p>关注类客户每月出具贷后监测报告，正常类客户按季度更新风险限额执行情况。</p></section>",
            "</body>",
            "</html>",
        ]
    )
    path.write_text("\n".join(html), encoding="utf-8")
    return path


def write_collateral_inventory_csv(path: Path) -> Path:
    """写入押品清单 CSV。"""
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.writer(file)
        writer.writerow(["押品名称", "押品类型", "评估价值", "抵质押率", "权属状态"])
        writer.writerow(["杭州厂区土地及厂房", "不动产", "2.4 亿元", "60%", "权属清晰"])
        writer.writerow(["数控加工中心设备组", "机器设备", "0.9 亿元", "40%", "已办理抵押登记"])
        writer.writerow(["核心客户应收账款池", "应收账款质押", "1.1 亿元", "50%", "需月度核验"])
        writer.writerow(["湖州仓储中心", "不动产", "0.8 亿元", "55%", "抵押顺位第一"])
        writer.writerow(["专利组合 HM-2026-A", "知识产权", "0.35 亿元", "20%", "需年度复评"])
        writer.writerow(["银行保证金账户", "保证金", "0.2 亿元", "100%", "已冻结"])
        writer.writerow(["出口订单应收款", "应收账款质押", "0.6 亿元", "45%", "需买方确认"])
    return path


def write_aml_due_diligence_txt(path: Path) -> Path:
    """写入反洗钱增强尽调记录。"""
    path.write_text(
        "\n".join(
            [
                "反洗钱增强尽调记录",
                "客户：华东智造集团",
                "尽调日期：2026-06-24",
                "",
                "一、触发原因",
                "客户新增异地大额回款账户，且 6 月 18 日出现同日多笔拆分回款。交易对手集中在宁波新能源装备有限公司、合肥智能产线科技有限公司和上海储能装备有限公司三家客户。",
                "",
                "二、补充材料",
                "客户已补充实际控制人声明、主要交易对手清单、资金来源证明、董事会授权文件、销售合同、发票、发货单和回款承诺函。",
                "",
                "三、复核工作底稿",
                "复核人员核对销售合同、银行流水、ERP 出库记录和发票开具时间，确认主要资金流向与订单交付进度基本匹配。异地账户开户原因与新能源客户集中回款安排一致，但仍需关注同日拆分回款是否用于规避内部审批阈值。",
                "",
                "四、风险判断",
                "本次交易背景具备合理商业目的，未发现空壳交易、明显循环资金或无真实贸易背景迹象。考虑客户集中度较高、逾期应收款占比超过管理线，建议维持中风险等级，并将异常回款账户纳入季度复核。",
                "",
                "五、名单筛查",
                "系统筛查结果显示，客户、实际控制人、主要董事和三家核心交易对手均未命中制裁名单、恐怖融资名单和高风险司法辖区名单。合规人员对交易对手工商登记、经营范围和历史合作记录进行了人工复核，未发现明显不一致。",
                "",
                "六、资金流向核验",
                "抽样核验 6 月 18 日至 6 月 28 日银行流水后发现，回款主要来自新能源设备订单，后续资金流向包括供应商预付款、税费缴纳和银行承兑保证金。供应商预付款与采购合同 PC-2026-0712 匹配，税费缴纳金额与 ERP 纳税申报计划基本一致。",
                "",
                "七、控制措施",
                "客户经理需将异地账户纳入重点监测账户清单，单笔超过 500 万元或同日累计超过 800 万元的入账需补充合同、发票和发货单。若出现交易对手突然变更、频繁跨行拆分、无合同背景大额入账或资金快速转出，应在一个工作日内提交异常交易复核单。",
                "",
                "人工复核结论依据包括：客户历史合作年限、订单毛利水平、发票开具节奏、物流签收记录、银行流水摘要和董事会融资决议。上述材料需在贷后档案中留存，便于后续抽查。",
                "",
                "八、后续动作",
                "客户经理需在每月贷后检查中补充核心客户回款流水，合规部在 2026-09-30 前完成二次复核。如出现新增对外担保、大额预付款或回款路径频繁变更，应重新启动增强尽调。",
            ]
        )
        + "\n",
        encoding="utf-8",
    )
    return path


def write_erp_snapshot_json(path: Path) -> Path:
    """写入可被 DoclingReader 直接解析的 ERP 财务快照 Docling JSON。"""
    from docling_core.types.doc import DoclingDocument

    doc = DoclingDocument(name="erp_financial_snapshot")
    doc.add_heading(text="ERP 财务快照：华东智造集团 2026H1", level=1)
    doc.add_text(
        label="text",
        text="\n".join(
            [
                "文件编号：ERP-FIN-2026H1-017",
                "数据来源：ERP 总账、应收模块、库存模块、银企直连流水核对表",
                "抽取时间：2026-06-30 22:15:00",
                "数据口径：合并范围内智能制造、新能源装备和工业软件三个事业部。",
                "2026H1 经营现金流为 0.8 亿元，同比下降 -57.9%，主要原因是核心客户回款周期延长、验收节点后移以及票据结算比例上升。",
            ]
        ),
    )
    doc.add_heading(text="一、经营概览", level=2)
    doc.add_text(
        label="text",
        text="\n".join(
            [
                "营业收入为 15.4 亿元，同比增长 20.3%；毛利率为 29.6%，较 2025H1 的 31.2% 下降 1.6 个百分点。",
                "经营现金流转化率为 5.2%，低于内部管理线 10%；短债覆盖倍数为 1.18x，低于银行约束底线 1.20x。",
                "逾期应收款占比为 6.2%，高于授信审查要求的 5% 上限；前五大客户收入占比为 48.5%，超过 45% 的关注阈值。",
                "存货周转天数为 104 天，较年初增加 12 天；呆滞库存占比为 8.7%，主要集中在视觉检测单元和高精度导轨备件。",
            ]
        ),
    )
    doc.add_heading(text="二、月度经营指标", level=2)
    doc.add_text(
        label="text",
        text="\n".join(
            [
                "月份 | 收入 | 经营现金流 | 应收回款 | 存货周转天数 | 逾期应收款占比",
                "2026-01 | 2.1 亿元 | 0.4 亿元 | 1.6 亿元 | 92 天 | 4.7%",
                "2026-02 | 1.8 亿元 | 0.1 亿元 | 1.2 亿元 | 95 天 | 5.1%",
                "2026-03 | 2.6 亿元 | 0.2 亿元 | 1.8 亿元 | 98 天 | 5.5%",
                "2026-04 | 2.9 亿元 | 0.0 亿元 | 1.9 亿元 | 101 天 | 5.9%",
                "2026-05 | 3.0 亿元 | -0.1 亿元 | 1.7 亿元 | 103 天 | 6.1%",
                "2026-06 | 3.0 亿元 | 0.2 亿元 | 2.0 亿元 | 104 天 | 6.2%",
            ]
        ),
    )
    doc.add_heading(text="三、现金流驱动项", level=2)
    doc.add_text(
        label="text",
        text="\n".join(
            [
                "应收账款增加占用现金 1.15 亿元，其中宁波新能源装备有限公司和合肥智能产线科技有限公司合计贡献 0.72 亿元。",
                "库存增加占用现金 0.42 亿元，主要为新能源产线扩产项目备货；供应商预付款增加占用现金 0.20 亿元。",
                "税费缴纳和银行承兑保证金合计流出 0.31 亿元；票据到期兑付和保函保证金占用需在三季度继续跟踪。",
                "若三季度核心客户回款延迟 30 天，预计短债覆盖倍数降至 1.07x；若延迟 60 天且库存继续上升，预计降至 0.96x。",
            ]
        ),
    )
    doc.add_heading(text="四、客户与应收账款风险", level=2)
    doc.add_text(
        label="text",
        text="\n".join(
            [
                "宁波新能源装备有限公司应收余额 3,260 万元，其中逾期 860 万元；回款责任人为周岚，需按周跟踪。",
                "合肥智能产线科技有限公司存在 180 天以上账龄，逾期金额 610 万元，风控动作是暂停新增赊销。",
                "上海储能装备有限公司 180 天以上逾期金额 410 万元，需补充买方确认函、发票清单和分期回款承诺。",
                "ERP 系统已将逾期应收款、核心客户收入占比和异常拆分回款纳入贷后预警清单。",
            ]
        ),
    )
    doc.add_heading(text="五、库存与采购联动", level=2)
    doc.add_text(
        label="text",
        text="\n".join(
            [
                "库存账面余额为 2.84 亿元，其中原材料 1.18 亿元、在产品 0.96 亿元、产成品 0.70 亿元。",
                "库龄超过 180 天的物料占比为 8.7%；视觉检测单元 VS-900、控制板卡 HM-CTRL-X7 和高精度导轨 HG-45A 需要专项复核。",
                "采购合同 PC-2026-0712 的付款条件为预付款 20%、验收后支付 50%、稳定运行 60 天后支付 30%。",
                "三季度采购排程应优先消化长库龄物料，新增采购必须绑定订单、发票、出库和回款计划。",
            ]
        ),
    )
    doc.add_heading(text="六、银行约束与贷后动作", level=2)
    doc.add_text(
        label="text",
        text="\n".join(
            [
                "银行约束要求短债覆盖倍数不低于 1.20x，当前短债覆盖倍数为 1.18x；要求逾期应收款占比不高于 5%，当前为 6.2%。",
                "新增提款需满足三季度经营现金流转正，且逾期应收款占比降至 5% 以下；提款用途限定为订单备货、核心客户交付和税费周转。",
                "客户经理需每月核验银行流水、ERP 出库记录、发票和回款计划；出现同日多笔拆分回款时需触发增强尽调。",
                "下一次 ERP 财务快照复核日期为 2026-09-30，复核重点为经营现金流、库存库龄、短债覆盖倍数和核心客户回款兑现情况。",
            ]
        ),
    )
    path.write_text(json.dumps(doc.export_to_dict(), ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def write_supplier_esg_pptx(path: Path) -> Path:
    """写入带明确布局的供应链 ESG 审查 PPTX。"""
    slides = [
        {
            "title": "供应链 ESG 审查结论需要绑定采购准入",
            "kicker": "华东智造集团 | 供应商：嘉兴精密部件有限公司 | 评级 B+ | 复核期：2026Q3",
            "sections": [
                (
                    "采购准入结论",
                    [
                        "维持合格供应商资格，但新增订单需绑定整改节点。",
                        "逾期未完成整改，暂停新增采购订单和预付款申请。",
                    ],
                ),
                (
                    "核心风险",
                    [
                        "废水在线监测未联网，危废转运台账留痕不足。",
                        "劳务派遣比例为 18%，高于 10% 的管理目标。",
                    ],
                ),
                (
                    "管理动作",
                    [
                        "供应链中心按月复核整改证据，法务合规部留存底稿。",
                        "付款审批与验收结论、整改进度和采购准入状态联动。",
                    ],
                ),
            ],
        },
        {
            "title": "环保合规缺口集中在监测联网和证据留痕",
            "kicker": "审查范围：环保合规、劳工用工、安全生产和包装材料回收",
            "sections": [
                (
                    "环保合规缺口",
                    [
                        "废水在线监测设备需在 2026-09-30 前完成联网。",
                        "危废转运联单需补齐批次编号、重量和接收单位回执。",
                    ],
                ),
                (
                    "用工与安全",
                    [
                        "劳务派遣比例需从 18% 降至 10% 以下。",
                        "夜班安全巡检记录缺少复核签字，需补充闭环记录。",
                    ],
                ),
                (
                    "包装回收",
                    [
                        "包装材料回收率当前 62%，目标提升至 75%。",
                        "新增订单优先使用可循环周转箱，减少一次性木箱使用。",
                    ],
                ),
            ],
        },
        {
            "title": "整改责任矩阵把 ESG 风险转成可跟踪动作",
            "kicker": "每一项整改都绑定责任部门、截止日、证据材料和采购动作",
            "sections": [
                (
                    "整改责任矩阵",
                    [
                        "环保联网：供应商设备部负责，2026-09-30 前提交联网截图和监测报告。",
                        "用工压降：供应商人事部负责，2026-08-31 前提交劳动合同和派遣比例台账。",
                    ],
                ),
                (
                    "证据材料",
                    [
                        "每月提交排污许可、危废联单、安全巡检、员工花名册和包装回收台账。",
                        "华东智造集团供应链中心抽查原始记录，不接受只截图不留底稿。",
                    ],
                ),
                (
                    "采购动作",
                    [
                        "完成整改前新增订单不得超过月均采购额的 70%。",
                        "连续两个季度 ESG 评分低于 B，启动替代供应商评估。",
                    ],
                ),
            ],
        },
        {
            "title": "采购联动措施明确触发条件和恢复路径",
            "kicker": "目标不是简单扣分，而是让供应风险进入订单、付款和供应商准入流程",
            "sections": [
                (
                    "暂停触发",
                    [
                        "逾期未完成环保联网或出现环保处罚，暂停新增采购订单。",
                        "出现重大安全事故，立即冻结验收和预付款流程。",
                    ],
                ),
                (
                    "恢复条件",
                    [
                        "整改证据通过复核后，恢复季度订单评审。",
                        "ESG 评分连续两个季度回升至 B+ 以上，可恢复常规采购比例。",
                    ],
                ),
                (
                    "贷后关联",
                    [
                        "核心供应商 ESG 问题会影响项目交付、库存周转和授信用途核验。",
                        "相关底稿同步进入贷后档案，供授信复核和采购审计调用。",
                    ],
                ),
            ],
        },
    ]
    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as deck:
        deck.writestr("[Content_Types].xml", _ppt_content_types(len(slides)))
        deck.writestr("_rels/.rels", _ppt_root_rels())
        deck.writestr("ppt/presentation.xml", _ppt_presentation_xml(len(slides)))
        deck.writestr("ppt/_rels/presentation.xml.rels", _ppt_presentation_rels(len(slides)))
        for index, slide in enumerate(slides, 1):
            deck.writestr(f"ppt/slides/slide{index}.xml", _ppt_slide_xml(slide, index))
    return path


def write_board_minutes_md(path: Path) -> Path:
    """写入董事会融资决议纪要 Markdown。"""
    path.write_text(
        """# 华东智造集团董事会融资决议纪要

会议日期：2026-06-28
会议地点：总部 18 层第一会议室
主持人：董事长 林启明
列席部门：财务中心、风控中心、法务合规部、供应链中心

## 一、会议背景

公司 2026 年上半年订单规模保持增长，但下游新能源客户回款周期延长，经营现金流较上年同期下降。财务中心提交了综合授信维持方案，风控中心同步提交了客户集中度、短债覆盖和逾期应收款专项分析。

## 二、决议事项

- 同意维持综合授信额度 3.5 亿元，用于流动资金、银行承兑汇票和保函。
- 授权财务中心在经营现金流转正后申请新增提款。
- 新增提款必须绑定订单、发票、出库和回款计划，不得用于关联拆借。
- 对外担保余额不得超过净资产的 25%，新增担保需经董事会三分之二以上通过。

## 三、讨论意见

风控中心认为，公司当前订单储备能够支撑未来两个季度收入确认，但现金回收明显滞后，应将授信使用与回款节点绑定。法务合规部提示，异地回款账户和同日拆分回款需要纳入增强尽调，相关材料应与贷后档案一并留存。

财务中心说明，三季度若核心客户按计划回款，经营现金流可恢复为正；若回款继续延迟 45 天以上，短债覆盖倍数可能低于 1.0x。董事会要求财务中心每两周更新资金缺口测算，并在出现提款需求前提交资金用途清单。

## 四、预算调整

三季度压缩非关键设备采购 1,200 万元，优先保障核心客户订单交付和逾期应收款清收。供应链中心需重新排定备货计划，避免长库龄物料继续增加。

## 五、责任分工

- 财务中心：负责授信提款、现金流预测、银行沟通和资金用途核验。
- 风控中心：负责应收账款账龄、短债覆盖倍数、押品价值和客户集中度监测。
- 法务合规部：负责增强尽调、合同条款复核和对外担保合规审查。
- 供应链中心：负责采购排程压降、供应商 ESG 整改跟踪和库存库龄治理。

## 六、风险跟踪

风控中心每月向董事会风险委员会报送应收账款账龄、短债覆盖倍数、客户集中度和银行流水异常摘要。若逾期应收款占比未在 2026-09-30 前降至 5% 以下，应冻结新增提款。

## 七、附件清单

1. 《2026H1 授信审查报告》
2. 《应收账款账龄与回款计划》
3. 《银行流水异常摘要》
4. 《采购合同付款条件说明》
5. 《增强尽调复核工作底稿》

## 八、表决结果

应到董事 7 人，实到董事 7 人。同意 7 票，反对 0 票，弃权 0 票。本决议自会议通过之日起生效。
""",
        encoding="utf-8",
    )
    return path


def write_bank_statement_scan(path: Path) -> Path:
    """写入模拟银行流水扫描图片。"""
    image = Image.new("RGB", PAGE_SIZE, "white")
    draw = ImageDraw.Draw(image)
    font_path = _font_path()
    title_font = ImageFont.truetype(str(font_path), 42)
    body_font = ImageFont.truetype(str(font_path), 27)
    small_font = ImageFont.truetype(str(font_path), 22)
    draw.rectangle((70, 70, 1170, 1520), outline="#334155", width=3)
    draw.text((100, 100), "银行流水扫描件", fill="#0f172a", font=title_font)
    draw.text((100, 165), "账户名称：华东智造集团    账号尾号：7821    币种：人民币", fill="#334155", font=small_font)
    headers = ["日期", "摘要", "借方", "贷方", "余额"]
    rows = [
        ["2026-06-18", "新能源客户回款", "", "5,000,000.00", "18,240,000.00"],
        ["2026-06-18", "新能源客户回款", "", "2,000,000.00", "20,240,000.00"],
        ["2026-06-22", "供应商预付款", "1,200,000.00", "", "19,040,000.00"],
        ["2026-06-26", "税费缴纳", "860,000.00", "", "18,180,000.00"],
        ["2026-06-28", "银行承兑保证金", "2,000,000.00", "", "16,180,000.00"],
    ]
    x_positions = [105, 285, 590, 770, 965]
    y = 245
    for x, header in zip(x_positions, headers):
        draw.text((x, y), header, fill="#0f172a", font=body_font)
    y += 54
    for row in rows:
        for x, value in zip(x_positions, row):
            draw.text((x, y), value, fill="#1f2937", font=small_font)
        y += 52
    draw.text((105, y + 38), "异常备注：同日多笔拆分回款，需与销售合同和发票核对。", fill="#991b1b", font=body_font)
    image.save(path)
    return path


def _new_pdf_page(doc: fitz.Document, subtitle: str) -> fitz.Page:
    """创建 PDF 页面并写入页眉。"""
    font_file = str(_font_path())
    page = doc.new_page(width=595, height=842)
    page.draw_rect(fitz.Rect(42, 30, 553, 80), color=(0.75, 0.82, 0.9), fill=(0.93, 0.96, 0.99))
    page.insert_text(fitz.Point(54, 52), CREDIT_REPORT_TITLE, fontname="msyh", fontfile=font_file, fontsize=15)
    page.insert_text(
        fitz.Point(54, 72),
        subtitle,
        fontname="msyh",
        fontfile=font_file,
        fontsize=8.5,
        color=(0.35, 0.4, 0.45),
    )
    return page


def _pdf_text(page: fitz.Page, text: str, x: float, y: float, *, fontsize: float, bold: bool = False) -> float:
    """按固定宽度写入 PDF 文本，并返回下一行纵坐标。"""
    font_file = str(_font_path())
    max_chars = 30 if bold else 36
    color = (0.04, 0.18, 0.32) if bold else (0.08, 0.12, 0.18)
    for line in _wrap_cjk(text, max_chars):
        page.insert_text(
            fitz.Point(x, y + fontsize + 2),
            line,
            fontname="msyh",
            fontfile=font_file,
            fontsize=fontsize,
            color=color,
        )
        y += 23 if bold else 20
    return y + (2 if bold else 4)


def _pdf_table(
    page: fitz.Page,
    title: str,
    y: float,
    headers: list[str],
    rows: list[list[str]],
    widths: list[float],
) -> float:
    """绘制带表头、边框和自动行高的 PDF 表格。"""
    font_file = str(_font_path())
    y = _pdf_text(page, title, 50, y, fontsize=11.5, bold=True)
    x = 50.0
    line_height = 13.0
    border = (0.55, 0.63, 0.72)
    header_fill = (0.86, 0.92, 0.97)
    row_fill = (0.98, 0.99, 1.0)
    all_rows = [headers, *rows]
    for row_index, row in enumerate(all_rows):
        wrapped = [
            _wrap_cjk(value, max(5, int(width / 9.4)))
            for value, width in zip(row, widths)
        ]
        row_height = max(30.0, max(len(lines) for lines in wrapped) * line_height + 14.0)
        fill = header_fill if row_index == 0 else row_fill
        x_cursor = x
        for cell_lines, width in zip(wrapped, widths):
            rect = fitz.Rect(x_cursor, y, x_cursor + width, y + row_height)
            page.draw_rect(rect, color=border, fill=fill, width=0.6)
            text = "\n".join(cell_lines)
            page.insert_textbox(
                fitz.Rect(x_cursor + 5, y + 7, x_cursor + width - 5, y + row_height - 4),
                text,
                fontname="msyh",
                fontfile=font_file,
                fontsize=8.4 if row_index else 8.8,
                color=(0.07, 0.11, 0.18),
            )
            x_cursor += width
        y += row_height
    return y + 8


def _pdf_footer(doc: fitz.Document) -> None:
    """给 PDF 所有页面补充页脚和页码。"""
    font_file = str(_font_path())
    total = doc.page_count
    for index, page in enumerate(doc, 1):
        page.draw_line(fitz.Point(50, 792), fitz.Point(545, 792), color=(0.75, 0.8, 0.86), width=0.5)
        page.insert_textbox(
            fitz.Rect(50, 798, 545, 816),
            f"内部授信审查资料 | 第 {index}/{total} 页 | 仅用于 RAG 评测样例",
            fontname="msyh",
            fontfile=font_file,
            fontsize=7.5,
            color=(0.38, 0.43, 0.5),
        )


def _wrap_cjk(text: str, limit: int) -> list[str]:
    """按字符宽度近似折行，避免裁切并尽量保留数字、英文和百分比片段。"""
    value = str(text or "")
    tokens = re.findall(r"[A-Za-z0-9@%./+\-]+|[\u4e00-\u9fff]|[^\u4e00-\u9fffA-Za-z0-9]+", value)
    lines: list[str] = []
    current = ""
    current_width = 0.0
    for token in tokens:
        token_width = _display_width(token)
        if current and current_width + token_width > limit:
            lines.append(current.rstrip())
            current = token.lstrip()
            current_width = _display_width(current)
        else:
            current += token
            current_width += token_width
    if current:
        lines.append(current.rstrip())
    return lines or [""]


def _display_width(value: str) -> float:
    """估算中英文混排文本的显示宽度。"""
    return sum(0.55 if ord(char) < 128 else 1.0 for char in value)


def _write_docx_package(path: Path, document_xml: str) -> None:
    """写入最小但带样式的 DOCX 包。"""
    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as docx:
        docx.writestr("[Content_Types].xml", _docx_content_types())
        docx.writestr("_rels/.rels", _docx_root_rels())
        docx.writestr("word/_rels/document.xml.rels", _docx_document_rels())
        docx.writestr("word/styles.xml", _docx_styles_xml())
        docx.writestr("word/document.xml", document_xml)


def _docx_paragraph(text: str, *, style: str | None = None) -> str:
    """生成 DOCX 段落 XML。"""
    spacing = '<w:spacing w:after="120" w:line="300" w:lineRule="auto"/>'
    style_xml = f'<w:pStyle w:val="{style}"/>' if style else ""
    return f"<w:p><w:pPr>{style_xml}{spacing}</w:pPr><w:r><w:t>{xml_escape(text)}</w:t></w:r></w:p>"


def _docx_table(headers: list[str], rows: list[list[str]]) -> str:
    """生成带列宽和单元格内边距的 DOCX 表格 XML。"""
    widths = [2600, 2600, 1800, 2600]
    grid = "".join(f'<w:gridCol w:w="{width}"/>' for width in widths)
    props = (
        '<w:tblPr><w:tblW w:w="9600" w:type="dxa"/>'
        '<w:tblCellMar><w:top w:w="120" w:type="dxa"/><w:left w:w="120" w:type="dxa"/>'
        '<w:bottom w:w="120" w:type="dxa"/><w:right w:w="120" w:type="dxa"/></w:tblCellMar></w:tblPr>'
    )
    table_rows = [_docx_table_row(headers, widths, header=True), *[_docx_table_row(row, widths) for row in rows]]
    return f"<w:tbl>{props}<w:tblGrid>{grid}</w:tblGrid>{''.join(table_rows)}</w:tbl>"


def _docx_table_row(values: list[str], widths: list[int], *, header: bool = False) -> str:
    """生成 DOCX 表格行 XML。"""
    cells = []
    for value, width in zip(values, widths):
        shade = '<w:shd w:fill="D9EAF7"/>' if header else ""
        bold = "<w:b/>" if header else ""
        cells.append(
            f'<w:tc><w:tcPr><w:tcW w:w="{width}" w:type="dxa"/>{shade}</w:tcPr>'
            f'<w:p><w:r><w:rPr>{bold}</w:rPr><w:t>{xml_escape(value)}</w:t></w:r></w:p></w:tc>'
        )
    return f"<w:tr>{''.join(cells)}</w:tr>"


def _docx_content_types() -> str:
    return """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>
  <Override PartName="/word/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.styles+xml"/>
</Types>"""


def _docx_root_rels() -> str:
    return """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>
</Relationships>"""


def _docx_document_rels() -> str:
    return """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rIdStyles" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>
</Relationships>"""


def _docx_styles_xml() -> str:
    return """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<w:styles xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">
  <w:style w:type="paragraph" w:styleId="Normal"><w:name w:val="Normal"/><w:rPr><w:rFonts w:ascii="Microsoft YaHei" w:eastAsia="Microsoft YaHei"/><w:sz w:val="21"/></w:rPr></w:style>
  <w:style w:type="paragraph" w:styleId="Heading1"><w:name w:val="heading 1"/><w:rPr><w:b/><w:color w:val="12324A"/><w:sz w:val="28"/></w:rPr></w:style>
</w:styles>"""


def _ppt_content_types(slide_count: int) -> str:
    overrides = "\n".join(
        f'  <Override PartName="/ppt/slides/slide{index}.xml" ContentType="application/vnd.openxmlformats-officedocument.presentationml.slide+xml"/>'
        for index in range(1, slide_count + 1)
    )
    return f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/ppt/presentation.xml" ContentType="application/vnd.openxmlformats-officedocument.presentationml.presentation.main+xml"/>
{overrides}
</Types>"""


def _ppt_root_rels() -> str:
    return """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="ppt/presentation.xml"/>
</Relationships>"""


def _ppt_presentation_xml(slide_count: int) -> str:
    slides = "\n".join(
        f'    <p:sldId id="{255 + index}" r:id="rId{index}"/>'
        for index in range(1, slide_count + 1)
    )
    return f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<p:presentation xmlns:p="http://schemas.openxmlformats.org/presentationml/2006/main" xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <p:sldIdLst>
{slides}
  </p:sldIdLst>
  <p:sldSz cx="{PPT_SLIDE_CX}" cy="{PPT_SLIDE_CY}" type="screen16x9"/>
  <p:notesSz cx="{PPT_SLIDE_CY}" cy="{EMU_PER_INCH * 10}"/>
  <p:defaultTextStyle>
    <a:defPPr><a:defRPr lang="zh-CN" sz="1800"><a:latin typeface="Microsoft YaHei"/><a:ea typeface="Microsoft YaHei"/></a:defRPr></a:defPPr>
  </p:defaultTextStyle>
</p:presentation>"""


def _ppt_presentation_rels(slide_count: int) -> str:
    rels = "\n".join(
        f'  <Relationship Id="rId{index}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/slide" Target="slides/slide{index}.xml"/>'
        for index in range(1, slide_count + 1)
    )
    return f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
{rels}
</Relationships>"""


def _ppt_slide_xml(slide: dict[str, object], slide_number: int) -> str:
    """生成带固定坐标和字号的 PPTX 幻灯片 XML。"""
    sections = list(slide["sections"])
    boxes = [
        _ppt_text_box(
            2,
            "Title",
            [str(slide["title"])],
            _emu(0.62),
            _emu(0.34),
            _emu(12.05),
            _emu(0.58),
            3600,
            bold=True,
            color="102A43",
        ),
        _ppt_text_box(
            3,
            "Kicker",
            [str(slide["kicker"])],
            _emu(0.64),
            _emu(0.98),
            _emu(11.9),
            _emu(0.36),
            1600,
            color="486581",
        ),
    ]
    positions = [(0.72, 1.72), (4.83, 1.72), (8.94, 1.72)]
    shape_id = 4
    for (heading, lines), (left, top) in zip(sections, positions):
        boxes.append(
            _ppt_text_box(
                shape_id,
                f"Section {shape_id} Heading",
                [str(heading)],
                _emu(left),
                _emu(top),
                _emu(3.42),
                _emu(0.44),
                2200,
                bold=True,
                color="0B3D5C",
            )
        )
        shape_id += 1
        boxes.append(
            _ppt_text_box(
                shape_id,
                f"Section {shape_id} Body",
                [f"- {line}" for line in lines],
                _emu(left),
                _emu(top + 0.52),
                _emu(3.42),
                _emu(4.25),
                1700,
                color="1F2937",
            )
        )
        shape_id += 1
    boxes.append(
        _ppt_text_box(
            shape_id,
            "Footer",
            [f"供应链 ESG 审查 | 第 {slide_number} 页"],
            _emu(0.64),
            _emu(7.02),
            _emu(11.9),
            _emu(0.24),
            1050,
            color="6B7280",
        )
    )
    body = "\n".join(boxes)
    return f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<p:sld xmlns:p="http://schemas.openxmlformats.org/presentationml/2006/main" xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">
  <p:cSld>
    <p:bg><p:bgPr><a:solidFill><a:srgbClr val="F8FAFC"/></a:solidFill></p:bgPr></p:bg>
    <p:spTree>
    <p:nvGrpSpPr><p:cNvPr id="1" name=""/><p:cNvGrpSpPr/><p:nvPr/></p:nvGrpSpPr><p:grpSpPr/>
{body}
    </p:spTree>
  </p:cSld>
  <p:clrMapOvr><a:masterClrMapping/></p:clrMapOvr>
</p:sld>"""


def _ppt_text_box(
    shape_id: int,
    name: str,
    paragraphs: list[str],
    x: int,
    y: int,
    cx: int,
    cy: int,
    font_size: int,
    *,
    bold: bool = False,
    color: str = "1F2937",
) -> str:
    """生成一个带固定坐标的 PPT 文本框。"""
    paragraph_xml = "\n".join(_ppt_paragraph(text, font_size, bold=bold, color=color) for text in paragraphs)
    return f"""    <p:sp>
      <p:nvSpPr><p:cNvPr id="{shape_id}" name="{xml_escape(name)}"/><p:cNvSpPr txBox="1"/><p:nvPr/></p:nvSpPr>
      <p:spPr>
        <a:xfrm><a:off x="{x}" y="{y}"/><a:ext cx="{cx}" cy="{cy}"/></a:xfrm>
        <a:prstGeom prst="rect"><a:avLst/></a:prstGeom>
        <a:noFill/><a:ln><a:noFill/></a:ln>
      </p:spPr>
      <p:txBody><a:bodyPr wrap="square" anchor="t"><a:spAutoFit/></a:bodyPr><a:lstStyle/>
{paragraph_xml}
      </p:txBody>
    </p:sp>"""


def _ppt_paragraph(text: str, font_size: int, *, bold: bool, color: str) -> str:
    """生成 PPT 文本段落 XML。"""
    bold_attr = ' b="1"' if bold else ""
    return (
        "        <a:p><a:r>"
        f'<a:rPr lang="zh-CN" sz="{font_size}"{bold_attr}>'
        f'<a:solidFill><a:srgbClr val="{color}"/></a:solidFill>'
        '<a:latin typeface="Microsoft YaHei"/><a:ea typeface="Microsoft YaHei"/>'
        f'</a:rPr><a:t xml:space="preserve">{xml_escape(text)}</a:t>'
        "</a:r></a:p>"
    )


def _emu(inches: float) -> int:
    """把英寸转换为 PPT 使用的 EMU 单位。"""
    return round(inches * EMU_PER_INCH)


def _font_path() -> Path:
    """查找本机可用中文字体。"""
    for path in FONT_CANDIDATES:
        if path.exists():
            return path
    raise RuntimeError("未找到可用字体，无法生成中文演示文档")


if __name__ == "__main__":
    raise SystemExit(main())
