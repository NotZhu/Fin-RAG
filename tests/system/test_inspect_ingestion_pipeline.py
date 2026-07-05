import json
import sys
import zipfile
from types import ModuleType

import fitz
from openpyxl import load_workbook
from PIL import Image
from llama_index.core import Document

from scripts import generate_demo_documents, inspect_ingestion_pipeline


class FakePreviewNode:
    def __init__(self, node_id, level, text, metadata=None, parent=None):
        self.node_id = node_id
        self.text = text
        self.parent_node = parent
        self.child_nodes = []
        self.metadata = {
            "chunk_id": node_id,
            "chunk_level": level,
            "parent_chunk_id": parent.node_id if parent else node_id,
            "root_chunk_id": parent.metadata["root_chunk_id"] if parent else node_id,
            "chunk_idx": 0,
            **(metadata or {}),
        }


def _install_empty_docling_modules(monkeypatch):
    fake_loader_module = ModuleType("finrag.ingestion.docling_loader")
    fake_loader_module.load_docling_documents = lambda *args, **kwargs: []

    fake_nodes_module = ModuleType("finrag.indexing.nodes")

    class UnusedDataPreparationModule:
        def __init__(self, *args, **kwargs):
            raise AssertionError("empty Docling results should stop before node building")

    fake_nodes_module.DataPreparationModule = UnusedDataPreparationModule
    monkeypatch.setitem(sys.modules, "finrag.ingestion.docling_loader", fake_loader_module)
    monkeypatch.setitem(sys.modules, "finrag.indexing.nodes", fake_nodes_module)


def _install_preview_modules(monkeypatch):
    fake_loader_module = ModuleType("finrag.ingestion.docling_loader")

    def load_docling_documents(path, *, knowledge_base_id, data_root=None):
        source = path.resolve()
        return [
            Document(
                text=f"Docling JSON for {source.name}",
                metadata={
                    "document_id": f"doc-{source.stem}",
                    "knowledge_base_id": knowledge_base_id,
                    "source_path": str(source),
                    "filename": source.name,
                    "file_type": source.suffix.lower().lstrip("."),
                    "parser_name": "docling",
                },
            )
        ]

    fake_loader_module.load_docling_documents = load_docling_documents

    fake_nodes_module = ModuleType("finrag.indexing.nodes")

    class FakeDataPreparationModule:
        def __init__(self, data_path, *, knowledge_base_id):
            self.data_path = data_path
            self.knowledge_base_id = knowledge_base_id

        def _build_hierarchical_nodes(self, documents):
            document = documents[0]
            metadata = dict(document.metadata)
            filename = metadata["filename"]
            root = FakePreviewNode(
                f"{filename}-root",
                1,
                f"{filename} 文档根节点",
                metadata,
            )
            section = FakePreviewNode(
                f"{filename}-section",
                2,
                f"{filename} 授信审批章节",
                metadata | {"section_title": "授信审批"},
                parent=root,
            )
            leaf = FakePreviewNode(
                f"{filename}-leaf",
                3,
                f"{filename} 叶子节点正文",
                metadata | {"section_title": "授信审批", "page_number": 1},
                parent=section,
            )
            root.child_nodes = [section]
            section.child_nodes = [leaf]
            return [root, section, leaf], [leaf]

    fake_nodes_module.DataPreparationModule = FakeDataPreparationModule
    monkeypatch.setitem(sys.modules, "finrag.ingestion.docling_loader", fake_loader_module)
    monkeypatch.setitem(sys.modules, "finrag.indexing.nodes", fake_nodes_module)


def test_inspect_pipeline_reports_empty_docling_result_without_success(monkeypatch, tmp_path, capsys):
    source = tmp_path / "broken.pdf"
    source.write_bytes(b"%PDF mocked")
    out_dir = tmp_path / "preview"

    _install_empty_docling_modules(monkeypatch)
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "inspect_ingestion_pipeline.py",
            str(source),
            "--out",
            str(out_dir),
        ],
    )

    exit_code = inspect_ingestion_pipeline.main()

    captured = capsys.readouterr()
    assert exit_code == 1
    assert "正在使用 Docling 解析" in captured.out
    assert "解析预览已生成" not in captured.out
    assert "Docling 未解析出任何 Document" in captured.err
    assert not (out_dir / "00_summary.json").exists()


def test_inspect_pipeline_uses_demo_documents_dir_when_source_is_omitted(monkeypatch, tmp_path, capsys):
    project_root = tmp_path / "project"
    demo_dir = project_root / "output" / "demo-documents"
    demo_dir.mkdir(parents=True)
    (demo_dir / "board_financing_minutes.md").write_text("# mocked", encoding="utf-8")

    monkeypatch.setattr(inspect_ingestion_pipeline, "PROJECT_ROOT", project_root)
    monkeypatch.setattr(inspect_ingestion_pipeline, "SRC_ROOT", project_root / "src")
    _install_empty_docling_modules(monkeypatch)
    monkeypatch.setattr(sys, "argv", ["inspect_ingestion_pipeline.py"])

    exit_code = inspect_ingestion_pipeline.main()

    captured = capsys.readouterr()
    assert exit_code == 1
    assert "未提供 source，默认使用" in captured.out
    assert str(demo_dir.resolve()) in captured.out
    assert "Docling 未解析出任何 Document" in captured.err


def test_inspect_pipeline_writes_readable_three_level_preview_for_each_source_file(monkeypatch, tmp_path):
    source_dir = tmp_path / "documents"
    source_dir.mkdir()
    (source_dir / "credit.md").write_text("# 授信审批\n正文", encoding="utf-8")
    (source_dir / "risk.txt").write_text("风险监测正文", encoding="utf-8")
    out_dir = tmp_path / "preview"

    _install_preview_modules(monkeypatch)
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "inspect_ingestion_pipeline.py",
            str(source_dir),
            "--out",
            str(out_dir),
            "--knowledge-base-id",
            "kb-test",
        ],
    )

    exit_code = inspect_ingestion_pipeline.main()

    assert exit_code == 0
    summary = json.loads((out_dir / "00_summary.json").read_text(encoding="utf-8"))
    assert summary["files_total"] == 2
    assert summary["files_parsed"] == 2
    assert summary["documents"] == 2
    assert summary["root_nodes"] == 2
    assert summary["section_nodes"] == 2
    assert summary["indexed_leaf_nodes"] == 2

    document_dirs = sorted((out_dir / "01_documents").iterdir())
    assert [path.name for path in document_dirs] == ["001_credit", "002_risk"]

    credit_tree = (document_dirs[0] / "02_three_level_tree.md").read_text(encoding="utf-8")
    assert "# credit.md" in credit_tree
    assert "L1 root" in credit_tree
    assert "L2 section" in credit_tree
    assert "L3 leaf" in credit_tree
    assert "授信审批" in credit_tree
    assert "credit.md 叶子节点正文" in credit_tree

    index = (out_dir / "00_index.md").read_text(encoding="utf-8")
    assert "credit.md" in index
    assert "risk.txt" in index
    assert "001_credit/02_three_level_tree.md" in index


def test_demo_generator_writes_business_document_formats(monkeypatch, tmp_path):
    out_dir = tmp_path / "demo-documents"
    monkeypatch.setattr(
        generate_demo_documents,
        "write_credit_review_pdf",
        lambda output: output.write_bytes(b"%PDF mocked"),
        raising=False,
    )
    monkeypatch.setattr(
        sys,
        "argv",
        ["generate_demo_documents.py", "--out-dir", str(out_dir), "--clean"],
    )

    exit_code = generate_demo_documents.main()

    assert exit_code == 0
    expected_files = {
        "credit_review_report.pdf",
        "accounts_receivable_aging.xlsx",
        "procurement_contract.docx",
        "risk_limit_policy.html",
        "collateral_inventory.csv",
        "aml_enhanced_due_diligence.txt",
        "erp_financial_snapshot.json",
        "supplier_esg_review.pptx",
        "board_financing_minutes.md",
        "bank_statement_scan.png",
    }
    assert {path.name for path in out_dir.iterdir() if path.is_file()} == expected_files
    assert not any(path.name.startswith("finrag_docling_demo") for path in out_dir.iterdir())

    from docling_core.types.doc import DoclingDocument

    erp_text = (out_dir / "erp_financial_snapshot.json").read_text(encoding="utf-8")
    erp_json = json.loads(erp_text)
    assert erp_json["schema_name"] == "DoclingDocument"
    erp_doc = DoclingDocument.model_validate_json(erp_text)
    exported_text = "\n".join(item.text for item in erp_doc.texts)
    assert "华东智造集团" in exported_text
    assert "2026H1 经营现金流为 0.8 亿元" in exported_text
    assert "短债覆盖倍数为 1.18x" in exported_text

    with zipfile.ZipFile(out_dir / "procurement_contract.docx") as docx:
        names = set(docx.namelist())
        document_xml = docx.read("word/document.xml").decode("utf-8")
        styles_xml = docx.read("word/styles.xml").decode("utf-8")
        rels_xml = docx.read("word/_rels/document.xml.rels").decode("utf-8")
    assert "word/styles.xml" in names
    assert 'w:styleId="Heading1"' in styles_xml
    assert 'Target="styles.xml"' in rels_xml
    assert "采购合同" in document_xml
    assert "付款条件" in document_xml
    assert document_xml.count('w:pStyle w:val="Heading1"') >= 8
    assert document_xml.count("<w:tbl>") >= 3
    assert "<w:tblCellMar>" in document_xml
    assert "<w:tblGrid>" in document_xml
    assert 'w:tcW w:w="' in document_xml

    with zipfile.ZipFile(out_dir / "accounts_receivable_aging.xlsx") as workbook:
        workbook_names = set(workbook.namelist())
        workbook_xml = "\n".join(
            workbook.read(name).decode("utf-8", errors="ignore")
            for name in workbook_names
            if name.endswith(".xml")
        )
    assert "xl/worksheets/sheet1.xml" in workbook_names
    assert "xl/worksheets/sheet2.xml" in workbook_names
    assert "xl/worksheets/sheet3.xml" in workbook_names
    assert "账龄明细" in workbook_xml
    assert "回款计划" in workbook_xml
    assert "集中度分析" in workbook_xml
    workbook = load_workbook(out_dir / "accounts_receivable_aging.xlsx", read_only=True, data_only=True)
    assert workbook["账龄明细"]["A1"].value == "客户名称"
    assert workbook["账龄明细"].max_row >= 9
    assert list(workbook["账龄明细"].iter_rows(min_row=1, max_row=1, values_only=True))[0] == (
        "客户名称",
        "账龄区间",
        "应收余额",
        "逾期金额",
        "回款责任人",
    )

    with zipfile.ZipFile(out_dir / "supplier_esg_review.pptx") as deck:
        deck_names = set(deck.namelist())
        presentation_xml = deck.read("ppt/presentation.xml").decode("utf-8")
        slide_text = "\n".join(
            deck.read(f"ppt/slides/slide{index}.xml").decode("utf-8") for index in range(1, 5)
        )
    assert "ppt/slides/slide1.xml" in deck_names
    assert "ppt/slides/slide2.xml" in deck_names
    assert "ppt/slides/slide3.xml" in deck_names
    assert "ppt/slides/slide4.xml" in deck_names
    assert 'cx="12192000"' in presentation_xml
    assert 'cy="6858000"' in presentation_xml
    assert slide_text.count("<p:spPr>") >= 9
    assert "环保合规缺口" in slide_text
    assert "采购准入结论" in slide_text
    assert "整改责任矩阵" in slide_text
    assert "华东智造集团" in slide_text
    assert "供应链 ESG 审查" in slide_text

    html_text = (out_dir / "risk_limit_policy.html").read_text(encoding="utf-8")
    assert html_text.count("<section") >= 4
    assert "贷后触发器" in html_text

    aml_text = (out_dir / "aml_enhanced_due_diligence.txt").read_text(encoding="utf-8")
    assert len(aml_text) > 900
    assert "复核工作底稿" in aml_text

    minutes_text = (out_dir / "board_financing_minutes.md").read_text(encoding="utf-8")
    assert len(minutes_text) > 1000
    assert "表决结果" in minutes_text

    with (out_dir / "collateral_inventory.csv").open(encoding="utf-8-sig") as file:
        assert sum(1 for _ in file) >= 8

    with Image.open(out_dir / "bank_statement_scan.png") as image:
        assert image.size == generate_demo_documents.PAGE_SIZE


def test_demo_pdf_contains_extractable_text(tmp_path):
    pdf_path = tmp_path / "demo.pdf"

    generate_demo_documents.write_credit_review_pdf(pdf_path)

    with fitz.open(pdf_path) as pdf:
        assert pdf.page_count >= 4
        extracted_text = "\n".join(page.get_text() for page in pdf)
        assert all(len(page.get_text().strip()) > 350 for page in pdf)
    normalized_text = " ".join(extracted_text.split())
    normalized_title = " ".join(generate_demo_documents.CREDIT_REPORT_TITLE.split())
    assert normalized_title in normalized_text
    assert "一、授信结论" in extracted_text
    assert "五、贷后监测安排" in extracted_text
    assert "财务指标矩阵" in extracted_text
    assert "压力测试情景" in extracted_text
    assert "贷后触发器矩阵" in extracted_text
